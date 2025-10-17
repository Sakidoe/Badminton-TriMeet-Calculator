import json
import openpyxl

# -> addEvent(Meet) :: add event categories to each team.
def addEvent(Meet): 
    for team in Meet:
        Meet[team]["MD"] = {}
        Meet[team]["MS"] = {}
        Meet[team]["XD"] = {}
        Meet[team]["WS"] = {}
        Meet[team]["WD"] = {}

# -> addPlayer() :: load players from Excel and insert into Meet JSON structure.
def addPlayer(team_dict, team_name):
    file_name = input(f"Enter Excel filename for {team_name} (include .xlsx): ").strip()
    try:
        wb = openpyxl.load_workbook(file_name)
        sh = wb.active
    except Exception as e:
        print(f"Error loading file: {e}")
        return

    data = []
    for i in range(1, sh.max_row + 1):
        for j in range(1, 4):
            val = sh.cell(row=i, column=j).value
            if val is not None:
                data.append(val)

    curr_event = None
    curr_rank = None

    for item in data:
        if item in ["MD", "MS", "XD", "WS", "WD"]:
            curr_event = item
        elif isinstance(item, (int, float)):
            curr_rank = f"Rank {int(item)}"
        elif isinstance(item, str):
            if curr_event is None or curr_rank is None:
                continue
            if curr_rank not in team_dict[curr_event]:
                team_dict[curr_event][curr_rank] = {"Player Name": [item]}
            else:
                team_dict[curr_event][curr_rank]["Player Name"].append(item)

    print(f"✅ Added players for team {team_name} successfully!")

# -> menu
def menu():
    print("----------------------Menu----------------------")
    print(" [A] Add Player Roster")
    print(" [V] View Roster")
    print(" [S] Save Changes")
    print(" [X] Exit")
    return input("\n<Your Choice>: ").strip().upper()

# -> save JSON to file
def save(Meet):
    json_format = json.dumps(Meet, indent=3)
    with open("save.json", "w") as f:
        f.write(json_format)
    print("✅ Saved to save.json")

# ------------------- MAIN PROGRAM -------------------

print("// PROGRAM START \\\\")
team_count = int(input("How many teams are there in this meet? (Max 3): ").strip())

while team_count < 2 or team_count > 3:
    team_count = int(input("❌ Invalid input. Enter 2 or 3 teams only: ").strip())

Meet = {}
teams = []

for i in range(1, team_count + 1):
    name = input(f"Enter team name #{i}: ").strip().upper()
    Meet[name] = {}
    teams.append(name)

addEvent(Meet)

if team_count == 2:
    print(f"Welcome to the meet between {teams[0]} and {teams[1]}!")
else:
    print(f"Welcome to the Trimeet between {', '.join(teams)}!")

while True:
    choice = menu()

    if choice == "A":
        team_input = input("Which team are you adding players for?: ").strip()
        team_key = team_input.upper().replace(".XLSX", "")

        if team_key in Meet:
            addPlayer(Meet[team_key], team_key)
        else:
            print(f"❌ Team '{team_input}' not found! Available teams: {list(Meet.keys())}")

    elif choice == "S":
        save(Meet)

    elif choice == "V":
        print(json.dumps(Meet, indent=3))

    elif choice == "X":
        print("Exiting program...")
        break

    else:
        print("❌ Invalid selection. Try again.")

print("Final Meet Data:")
print(json.dumps(Meet, indent=3))
